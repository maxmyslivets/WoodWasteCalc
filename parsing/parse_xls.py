import string
import re
from pathlib import Path
from openpyxl import load_workbook

from wood.wood import Wood, Trunk
from errors.parse_errors import *

from config.config import Config

config = Config()

from taxation.species_db import Species

shrub_species, wood_species = Species().get_species()


class RawWood:
    """
    Класс для обработки данных о дереве из xls файла
    """

    def __init__(self, number: str, name: str, quantity: str, diameter: str, height: str) -> None:
        """
        :param number: Номер дерева
        :param name: Название дерева
        :param quantity: Количество деревьев или стволов
        :param diameter: Диаметр дерева
        :param height: Высота дерева
        """
        self._number = number
        self._name = name
        self._quantity = quantity
        self._diameter = diameter
        self._height = height

        # дополнительные переменные
        self.number: list[str]
        self.specie: str
        self.is_shrub: bool
        self.trunk_count: int
        self.quantity: int
        self.diameter: list[float]
        self.height: list[float]
        self.quantity_is_area: bool

    def __repr__(self) -> str:
        return f"'{self._number}'\t'{self._name}'\t'{self._quantity}'\t" \
               f"'{self._diameter}'\t'{self._height}'"

    def parse_number(self) -> None:
        """
        Форматирование номера дерева или деревьев в таблице
        """
        result = []
        parts = self._number.translate({ord(c): None for c in string.whitespace}).split(',')
        for part in parts:
            part = part.strip().replace(' ', '')
            if '-' in part:
                start, end = part.split('-')
                try:
                    try:
                        is_float = False
                        start = int(start)
                        end = int(end)
                        result.extend([str(_) for _ in range(int(start), int(end) + 1)])
                    except ValueError as e:
                        is_float = True
                        int_number_part = str(start.split('.')[0])
                        start = start.split('.')[1]
                        end = end.split('.')[1]
                        result.extend([int_number_part+'.'+str(_) for _ in range(int(start), int(end) + 1)])
                # WARNING опасный костыль !!!!!!!!!!!!!!!
                # работает только для номеров типа `32 -32а`, где подразумевается только два номера
                except AttributeError as e:
                    result.extend([start, end])
            else:
                result.append(part)
        self.number = result

    def parse_specie(self) -> None:
        """
        Парсинг строки наименования породы
        """

        # Название породы
        self.specie = self._name.split()[0].lower()

        # Определение: кустарник или дерево
        if self.specie in shrub_species or "(поросль)" in self._name:
            self.is_shrub = True
        elif self.specie in wood_species and "(поросль)" not in self._name:
            self.is_shrub = False
        else:
            raise ParseSpecieError("Не удалось определить кустарник или дерево")

        # Определение количества стволов
        match = re.search(r'(\d+)\s*ствол', self._name, re.IGNORECASE)
        self.trunk_count = int(match.group(1)) if match else 1

    def parse_quantity(self) -> None:
        """
        Определение количества деревьев или площади поросли
        """

        self.quantity_is_area = "м" in self._quantity

        quantity_match = re.search(r'\d+', self._quantity)
        self.quantity = int(quantity_match.group(0))

    def parse_diameter(self) -> None:
        """
        Форматирование списка диаметров
        """

        self._diameter = self._diameter.replace(' ', ',')

        if self.quantity == 1 and self._diameter.count(",") == 1 and self.trunk_count == 1:
            self.diameter = [float(self._diameter.replace(',', '.'))]
            return

        if ';' in self._diameter:
            self._diameter = self._diameter.replace(',', '.')
            self._diameter = self._diameter.replace(';', ',')

        self._diameter = self._diameter.replace(',,', ',')
        self._diameter = self._diameter.replace(',.', ',')

        result = []

        # Замена кириллического "х" на латинское "x"
        diameters = self._diameter.replace("х", "x").replace("Х", "x")

        # Разделение строки по запятым
        parts = diameters.split(',')

        for part in parts:
            part = part.strip()
            # Поиск шаблона "числоxчисло"
            match = re.match(r'(\d+(\.\d+)?)x(\d+)', part)
            if match:
                # Если найдено, распаковываем значения
                num, _, repeat = match.groups()
                result.extend([float(num)] * int(repeat))
            elif part.isdigit() or re.match(r'^\d+(\.\d+)?$', part):
                # Если часть является числом, добавляем его в результат
                result.append(float(part))
            elif part == "-":
                # Специальный случай для "-"
                result.append(2.0)

        if self.trunk_count > len(result):
            result *= self.trunk_count
        if len(result) < len(self.number):
            # if len(result) == 1 and not self.is_shrub:
            result *= len(self.number)
            # else:
            #     raise ParseDiameterError(f"Диаметров меньше чем номеров")
        if len(result) > len(self.number) and not self.quantity_is_area:
            if not self.trunk_count > len(self.number) == 1:
                if len(self.number) == 1 and len(self.number[0]) == len(self._number):
                    self.number *= len(result)
                else:
                    raise ParseDiameterError(f"Диаметров больше чем номеров")

        self.diameter = result

    def parse_height(self) -> None:
        """
        Форматирование списка высот
        """
        if all(word not in self._height for word in ["выше", "ниже", "до", "от"]):
            self._height = self._height.replace(' ', ',')

        if self.quantity == 1 and self._height.count(",") == 1 and self.trunk_count == 1:
            self.height = [float(self._height.replace(',', '.'))]
            return

        if ';' in self._height:
            self._height = self._height.replace(',', '.')
            self._height = self._height.replace(';', ',')

        self._height = self._height.replace(',,', ',')
        self._height = self._height.replace(',.', ',')

        result = []

        # Замена кириллического "х" на латинское "x"
        heights = self._height.replace("х", "x").replace("Х", "x")

        # Разделение строки по запятым
        parts = heights.split(',')

        for part in parts:
            part = part.strip().lower()
            # Поиск шаблона "числоxчисло"
            match = re.match(r'(\d+(\.\d+)?)x(\d+)', part)
            if match:
                # Если найдено, распаковываем значения
                num, _, repeat = match.groups()
                result.extend([float(num)] * int(repeat))
            elif part.isdigit() or re.match(r'^\d+(\.\d+)?$', part):
                # Если часть является числом, добавляем его в результат
                result.append(float(part))
            elif "свыше" in part or "от" in part:
                # Обработка случая "свыше Xм"
                number_match = re.search(r'\d+', part)
                if number_match:
                    result.append(float(number_match.group()) + 0.5)
            elif "ниже" in part or "до" in part:
                # Обработка случая "ниже Xм"
                number_match = re.search(r'\d+', part)
                if number_match:
                    h = float(number_match.group())
                    result.append(h - 0.5 if h > 0.5 else h)

        if self.trunk_count > len(result):
            result *= self.trunk_count
        if len(result) < len(self.number):
            # if len(result) == 1 and not self.is_shrub:
                result *= len(self.number)
            # else:
            #     raise ParseDiameterError(f"Высот меньше чем номеров")
        if len(result) > len(self.number) and not self.quantity_is_area:
            if not self.trunk_count > len(self.number) == 1:
                if len(self.number) == 1 and len(self.number[0]) == len(self._number):
                    self.number *= len(result)
                else:
                    raise ParseDiameterError(f"Высот больше чем номеров")

        self.height = result

    def parse(self) -> list[Wood]:
        """

        :return:
        """

        self.parse_number()
        self.parse_specie()
        self.parse_quantity()
        self.parse_diameter()
        self.parse_height()

        woods = []

        # try:
        for idx_wood in range(len(self.number)):
            trunks = []
            if self.trunk_count > 1:
                for idx_trunk in range(self.trunk_count):
                    trunks.append(Trunk(diameter=self.diameter[idx_trunk], height=self.height[idx_trunk]))
                woods.append(Wood(name=self._name, number=self.number[idx_wood], specie=self.specie,
                                  is_shrub=self.is_shrub, trunks=trunks))
            else:
                trunks.append(Trunk(diameter=self.diameter[idx_wood], height=self.height[idx_wood]))
                woods.append(Wood(name=self._name, number=self.number[idx_wood], specie=self.specie,
                                  is_shrub=self.is_shrub, trunks=trunks,
                                  area=self.quantity if self.quantity_is_area else None))
        # except Exception as e:
        #     import traceback
        #     print(traceback.format_exc())
        #     raise ParseError(f"Ошибка получения объекта дерева {self.__repr__()}. {e}")

        return woods


class XLSWoodParser:
    """
    Класс для чтения excel файлов

    1 строка входного файла должна представлять заголовки таблицы.
    """
    def __init__(self) -> None:

        self.columns = [
            config.table_structure.number,
            config.table_structure.specie,
            config.table_structure.quality,
            config.table_structure.diameter,
            config.table_structure.height
        ]

    def parse(self, filepath: Path) -> list:
        """
        Парсинг excel файла. Получение двумерного массива данных таблицы.
        :param filepath: Путь до файла XLS
        :return: двумерный массив данных таблицы
        """
        # Чтение исходного файла Excel
        wb = load_workbook(filepath)
        ws = wb.active

        raw_woods = []
        for row in range(2, ws.max_row + 1):
            wood_list = []
            for i, col in enumerate(self.columns, start=1):
                wood_list.append(str(ws[f'{col}{row}'].value))
            raw_woods.append(wood_list)

        return raw_woods

