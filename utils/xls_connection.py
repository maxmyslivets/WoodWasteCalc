import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from pathlib import Path

from openpyxl.utils import get_column_letter

from config.config import Config


config = Config()


class XLSConnection:
    """
    Класс для объединения полученных с помощью wood_waste_calc XLS файлов в один файл и подсчета отходов по каждой группе

    Алгоритм работы класса:
    - Читает все XLS файлы в директории.
    - Создает новый XLS файл, на первом листе которого будут содержаться суммы отходов по каждой группе
    - Копирует данные со всех файлов в отдельные листы итогового XLS файла
    - Добавляет на первый лист информацию: на каком листе данные с какого файла
        например:
            лист 1 - улица Молодежная
            лист 2 - улица Вербная
            лист 3 - общественная зона
    - Сохраняет XLS файл в директорию, переданную конструктору класса
    """
    def __init__(self, directory: Path) -> None:
        self.directory = directory
        self.files = [file for file in self.directory.iterdir() if file.is_file()]

    def create_out_xlsx(self) -> None:
        """
        Создает новый XLS файл, на первом листе которого будут содержаться суммы отходов по каждой группе
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Суммарные отходы"
        # соединяем ячейки для заголовков на первом листе
        self.ws.merge_cells('A1:A2')
        self.ws.merge_cells('B1:B2')
        self.ws.merge_cells('C1:E1')
        self.ws.merge_cells('F1:G1')
        # заголовки
        titles = [('A1', 'Имя листа'), ('B1', 'Наименование группы'), ('C1', 'Деревья'), ('F1', 'Кустарники'),
                  ('C2', 'Стволы'), ('D2', 'Сучья'), ('E2', 'Пни'), ('F2', 'Сучья'), ('G2', 'Пни')]
        # пишем заголовки и задаем стили
        for cell, title in titles:
            self.ws[cell].value = title
            self.ws[cell].font = Font(bold=True)
            self.ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    def add_info_about_input_files(self, sheet_name: str, group_name: str) -> None:
        """
        Добавляет на первый лист информацию: на каком листе данные с какого файла
        :param sheet_name: имя листа
        :param group_name: название группы
        """
        self.ws.append([sheet_name, group_name])

    def copy_data_to_out_xlsx(self) -> None:
        """
        Копирует данные со всех файлов в отдельные листы итогового XLS файла
        """
        for idx, file in enumerate([f for f in self.files if 'summary' not in f.name]):
            wb = load_workbook(file)
            ws = wb.active
            sheet_name = str(idx+1)
            self.wb.create_sheet(sheet_name)
            sheet = self.wb[sheet_name]
            for row in ws.iter_rows():
                sheet.append([cell.value for cell in row])
            group_name = Path(file).stem.replace("_out_shrub", " (кустарники)").replace(
                "_out_wood", " (деревья)")
            self.add_info_about_input_files(sheet_name, group_name)

    def calculation_summary(self) -> None:
        """
        Подсчет суммарных отходов по каждой группе
        """
        letters = ["C", "D", "E", "F", "G"]
        for idx, row in enumerate(list(self.ws)[2:]):
            num_row = str(idx+3)
            sheet_name = str(self.ws[f"A{num_row}"].value)
            if str(self.ws[f"B{num_row}"].value).endswith('(кустарники)'):
                self.ws[f"F{num_row}"].value = f"=SUM('{sheet_name}'!I:I)"
                self.ws[f"G{num_row}"].value = f"=SUM('{sheet_name}'!J:J)"
            if str(self.ws[f"B{num_row}"].value).endswith('(деревья)'):
                self.ws[f"C{num_row}"].value = f"=SUM('{sheet_name}'!H:H)"
                self.ws[f"D{num_row}"].value = f"=SUM('{sheet_name}'!I:I)"
                self.ws[f"E{num_row}"].value = f"=SUM('{sheet_name}'!J:J)"
            for letter in letters:
                self.ws[f"{letter}{num_row}"].number_format = '0.00000'
                self.ws[f"{letter}{num_row}"].alignment = Alignment(horizontal='center', vertical='center')
        num_row = self.ws.max_row + 1
        # подпись заголовков итоговых значений
        self.ws.merge_cells(f'C{num_row}:E{num_row}')
        self.ws.merge_cells(f'F{num_row}:G{num_row}')
        self.ws[f'C{num_row}'].value = 'ИТОГО ДЕРЕВЬЯ'
        self.ws[f'F{num_row}'].value = 'ИТОГО КУСТАРНИКИ'
        self.ws[f'C{num_row}'].font = Font(bold=True)
        self.ws[f'F{num_row}'].font = Font(bold=True)
        self.ws[f'C{num_row}'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws[f'F{num_row}'].alignment = Alignment(horizontal='center', vertical='center')
        # расчет итоговых значений
        for letter in letters:
            self.ws[f"{letter}{num_row+1}"].value = f"=SUM({letter}3:{letter}{num_row-1})"
            self.ws[f"{letter}{num_row+1}"].number_format = '0.00000'
            self.ws[f"{letter}{num_row+1}"].alignment = Alignment(horizontal='center', vertical='center')

    def styled_table(self) -> None:
        """
        Стилизует таблицу
        """
        # настройка автоподбора ширины ячеек
        for col in range(1, self.ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, self.ws.max_row + 1):
                try:
                    if len(str(self.ws[column_letter + str(row)].value)) > max_length:
                        max_length = len(str(self.ws[column_letter + str(row)].value))
                except:
                    continue
            self.ws.column_dimensions[column_letter].width = max_length + 5
        # выравнивание первого столбца
        for row in range(1, self.ws.max_row + 1):
            self.ws['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center')
        # рисование границ
        th = Side(border_style="thin", color="FF000000")
        bd = Side(border_style="medium", color="FF000000")
        for col in range(1, self.ws.max_column + 1):
            column_letter = get_column_letter(col)
            for row in range(1, self.ws.max_row + 1):
                self.ws[column_letter + str(row)].border = Border(left=th, top=th, right=th, bottom=th)
                if row == self.ws.max_row-1 and column_letter in ["C", "D", "E", "F", "G"]:
                    self.ws[column_letter + str(row)].border = Border(left=th, top=bd, right=th, bottom=th)

    def save_xlsx(self) -> None:
        """
        Сохраняет XLS файл в директорию, переданную конструктору класса
        """
        # сохраняем файл
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        self.wb.save(self.directory / f"summary_{timestamp}.xlsx")


def xls_connection():
    connector = XLSConnection(config.directories.out_directory)
    connector.create_out_xlsx()
    connector.copy_data_to_out_xlsx()
    connector.calculation_summary()
    connector.styled_table()
    connector.save_xlsx()


if __name__ == "__main__":
    xls_connection()
